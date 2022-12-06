from openpyxl import load_workbook

plan = load_workbook("t.xlsx")
ativa = plan.active
r = ativa.max_row
cells = ativa["A1": f"H{r}"]
t=0
for a, b, c, d, e, f, g, h in cells:
    close = f"{c.value}"
    open = f"{b.value}"
    sub = float(close) - float(open)
    if sub == 0:
        f.value = "D"
    elif sub < 0:
        f.value = "R"
    elif sub > 0:
        f.value = "G"
    t=t+1
    print(t)
    plan.save("t.xlsx")

print('done')
